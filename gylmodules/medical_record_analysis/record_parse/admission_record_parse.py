import xml.etree.ElementTree as ET
import re
import os

from gylmodules.medical_record_analysis.build_cda.cda_xml_data_build import assembling_cda_record

"""
====================================================================================================
============================================ 入院记录解析 ============================================
====================================================================================================
"""

section_info = {}

sign_set = set()


def parse_patient_document_by_str(xml_str):
    root = ET.fromstring(xml_str)
    # root = tree.getroot()
    patient_info = {}

    # 提取姓名，性别，年龄等基本信息
    patient_info['姓名'] = root.find(".//element[@title='姓名']").text
    patient_info['pat_no'] = root.find(".//element[@sid='6A67D9D88F06411096CFD9690C452186']").text
    patient_info['pat_bed'] = root.find(".//element[@sid='CEB7B32AE5C745BA8A800A93E9BBE5A5']").text
    patient_info['pat_dept'] = root.find(".//element[@sid='E0AACD7387FC43C2BF15B375250DAD3F']").text
    patient_info['入院时间'] = root.find(".//element[@title='入院日期']").text
    # 病史叙述者 有可能为空
    if root.find(".//e_enum[@title='病史陈述者']") is not None and root.find(".//e_enum[@title='病史陈述者']").find('enumvalues/element') is not None:
        patient_info['病史叙述者'] = root.find(".//e_enum[@title='病史陈述者']").find('enumvalues/element').text
    else:
        patient_info['病史叙述者'] = ''

    header = root.find('./document')

    # 特殊处理
    for utext in header.iter('utext'):
        if utext.text is None:
            continue
        text = utext.text.replace(' ', '').replace(':', '').replace('：', '')
        if text == '与患者关系':
            utest_no = int(utext.get('no'))
            value_no = str(utest_no + 1)
            patient_info['与患者关系'] = root.find(".//utext[@no=" + "\'" + value_no + "\'" "]").text.strip(": ")
        if text == '职业' and '职业' not in patient_info:
            utest_no = int(utext.get('no'))
            value_no = str(utest_no + 1)
            patient_info['职业'] = root.find(".//utext[@no=" + "\'" + value_no + "\'" "]").text.strip(": ")

    for section in header.iter('section'):
        parse_hpi(section, patient_info,  patient_info['姓名'] + '入院记录')

    return patient_info


def parse_patient_document(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    patient_info = {}

    # 提取姓名，性别，年龄等基本信息
    patient_info['姓名'] = root.find(".//element[@title='姓名']").text
    patient_info['pat_no'] = root.find(".//element[@sid='6A67D9D88F06411096CFD9690C452186']").text
    patient_info['pat_bed'] = root.find(".//element[@sid='CEB7B32AE5C745BA8A800A93E9BBE5A5']").text
    patient_info['pat_dept'] = root.find(".//element[@sid='E0AACD7387FC43C2BF15B375250DAD3F']").text
    patient_info['入院时间'] = root.find(".//element[@title='入院日期']").text
    # 病史叙述者 有可能为空
    if root.find(".//e_enum[@title='病史陈述者']") is not None and root.find(".//e_enum[@title='病史陈述者']").find('enumvalues/element') is not None:
        patient_info['病史叙述者'] = root.find(".//e_enum[@title='病史陈述者']").find('enumvalues/element').text
    else:
        patient_info['病史叙述者'] = ''

    header = root.find('./document')

    # 特殊处理
    for utext in header.iter('utext'):
        if utext.text is None:
            continue
        text = utext.text.replace(' ', '').replace(':', '').replace('：', '')
        if text == '与患者关系':
            utest_no = int(utext.get('no'))
            value_no = str(utest_no + 1)
            patient_info['与患者关系'] = root.find(".//utext[@no=" + "\'" + value_no + "\'" "]").text.strip(": ")
        if text == '职业' and '职业' not in patient_info:
            utest_no = int(utext.get('no'))
            value_no = str(utest_no + 1)
            patient_info['职业'] = root.find(".//utext[@no=" + "\'" + value_no + "\'" "]").text.strip(": ")

    for section in header.iter('section'):
        parse_hpi(section, patient_info, xml_file)

    return patient_info


def parse_enum(enum, rangexml: str = ''):
    if not rangexml:
        if enum.find('enumvalues'):
            rangexml = enum.find('enumvalues').tail if enum.find('enumvalues') else ''
        else:
            rangexml = enum.text if enum.text else ''
    if not rangexml:
        return ''
    # 使用正则表达式匹配并提取 </root>' 后面的汉字部分
    pattern = re.compile(r"rangexml='[^']*'\s*(\S+)")
    match = pattern.search(rangexml)
    if match:
        extracted_text = match.group(1)
        return extracted_text
    else:
        print(enum.get('title'), " ===> 未能提取出汉字部分")
    return ''


def skip_continue(father_sid, sid):
    if father_sid == 'AB115ABF651443FA847D6DB6BDB0153E' and sid == '1A49854542FC4788ADFD8BA0273F4E9A':
        # 婚育史 过滤部分无效数据
        return True
    if father_sid == '7F98425DF2AD41109727949ABE9B3773' and sid == 'D850CDB8B9DD490D87CC6F752131BCC7':
        # 初步诊断 过滤 “单机这里选择职称”
        return True
    if father_sid == '61E75BB1B4B346008DFBEC79BEDE630F' and (sid == '0C6CFAB22A2C4AADAA480DA199D738D1' or sid == '138B3F46D506445FAA0B38F35739E8AB'):
        # 专科情况 过滤 “报告卡编码” "有无"
        return True
    return False


# 解析病历
def parse_hpi(section, info_dict, file):
    try:
        sid = section.get('sid') if section.get('sid') else section.get('iid')
        title = section.get('title')
        value = ''
        value_dict = {}
        for child in section:
            tag = child.tag
            if tag == 'break' or (tag == 'utext' and (str(child.text).replace(' ', '') == ':' or str(child.text).replace(' ', '') == '：')):
                continue
            if tag == 'utext':
                if child.text and not child.text.__contains__('家属签字确认'):
                    value = value + child.text.strip() if child.text else ''
            elif tag == 'element':
                if skip_continue(section.get('sid'), child.get('sid')):
                   continue
                text = child.text if child.text else '/'
                if text.__contains__('textstyleno') and child.get('value'):
                    text = child.get('value')
                value = value + text
                # 婚育史 过滤部分无效数据
                # child_sid = child.get('sid') if child.get('sid') else child.get('iid')
                value_dict[child.get('title')] = text if 'value' not in child.attrib or not 'unit' in child.attrib else child.get('value', '') + ' ' + child.get('unit', '')
            elif tag == 'e_enum' or tag == 'e_list':
                if len(child) < 1:
                    continue
                if skip_continue(section.get('sid'), child.get('sid')) or child.get('title').replace(' ', '') == '报告卡编码' or child.get('title').replace(' ', '') == '单机这里选择职称':
                   continue
                text = parse_enum(child)
                value = value + text
                # 婚育史 过滤部分无效数据
                # child_sid = child.get('sid') if child.get('sid') else child.get('iid')
                value_dict[child.get('title')] = text
            elif tag == 'group':
                for group in child.findall('group'):
                    group_text = ''
                    for item in group:
                        if (item.tag == 'utext' or item.tag == 'element') and item.text:
                            group_text = group_text + item.text.replace("textstyleno=2 unitstr='", '')
                        elif item.tag == 'e_enum':
                            enumtext = parse_enum(item)
                            group_text = group_text + enumtext
                            value_dict[item.get('title')] = enumtext
                            # item_sid = item.get('sid') if item.get('sid') else item.get('iid')
                            # if item_sid:
                            #     value_dict[item_sid] = enumtext
                    value = value + group_text
                    # child_sid = child.get('sid') if child.get('sid') else child.get('iid')
                    value_dict[group.get('title')] = group_text
            elif tag == 'signature':
                signplaceholder = child.get('signplaceholder')
                if signplaceholder:
                    signplaceholder = signplaceholder.replace('[', '').replace('签名]', '')
                displayinfo = child.get('displayinfo')
                if displayinfo:
                    info_dict[signplaceholder] = displayinfo
                    sign_set.add((child.get('signplaceholder'), child.get('iid')))
            elif tag in ('tab', 'image', 'patisign', 'table'):
                # 暂时不处理，没意义
                continue
            else:
                if not (file.__contains__('病程记录') and tag == 'section'):
                    print('===> 未处理 ', title, tag)

        if len(value_dict) < 2:
            info_dict[title] = value.strip()
        else:
            value_dict['value'] = value.strip()
            info_dict[title] = value_dict
    except Exception as e:
        print(file, '===> parse_hpi 解析异常', e)


def write_to_excel(data):
    import pandas as pd
    from openpyxl import load_workbook

    # 假设你已有的Excel文件名为'existing_file.xlsx'
    # file_path = '/Users/gaoyanliang/nsyy/病历解析/入院记录/parse_data.xlsx'
    file_path = 'parsed_data.xlsx'
    # 将字典转换为DataFrame
    df = pd.DataFrame(data)
    # 尝试加载现有的Excel文件
    try:
        # 使用openpyxl加载工作簿
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        writer.book = book

        # 检查是否已有数据表
        if 'Sheet1' in book.sheetnames:
            existing_df = pd.read_excel(file_path, sheet_name='Sheet1', engine='openpyxl')
            combined_df = pd.concat([existing_df, df], ignore_index=True)
        else:
            combined_df = df

        # 将合并后的数据写入Excel文件
        combined_df.to_excel(writer, index=False, sheet_name='Sheet1')
        writer.save()
        writer.close()
    except FileNotFoundError:
        # 如果文件不存在，创建新文件
        df.to_excel(file_path, index=False, engine='openpyxl')


def merge_data(patient_info):
    # 一般状况检查
    general_survey = ''
    # 皮肤和粘膜检査
    skin_and_mucosal = ''
    # 全身浅表淋巴结检查
    node = ''
    # 头部检査
    head = ''
    # 颈部检査
    neck = ''
    # 腹部检査
    abdomen = ''
    # 胸部检査
    chest = ''
    # 肛门指诊检查
    digital_rectal = ''
    # 外生殖器检査
    external_genital = ''
    # 脊柱检查
    spinal = ''
    # 四肢检查
    limb = ''
    # 神经系统
    neurological = ''
    # 中医四诊
    chinese = ''

    for key, value in patient_info.items():
        if key is None:
            continue
        if key.replace(' ', '').__contains__('声音') \
                or key.replace(' ', '').__contains__('气味') \
                or key.replace(' ', '').__contains__('舌象') \
                or key.replace(' ', '').__contains__('脉象'):
            if type(value) == dict:
                chinese += key + ':' + value['value'] + ','
            else:
                chinese += key + ':' + value + ','

    if '体格检查' in patient_info:
        data = patient_info['体格检查']
        # 一般状况检查
        for key, value in data.items():
            if key is None:
                continue
            if key.__contains__('淋巴'):
                node += key + ':' + value + ','
            elif key.__contains__('皮肤') or key.__contains__('粘膜'):
                skin_and_mucosal += key + ':' + value + ','
            elif key.__contains__('颈'):
                neck += key + ':' + value + ','
            elif key.__contains__('头'):
                head += key + ':' + value + ','
            elif key.__contains__('腹'):
                abdomen += key + ':' + value + ','
            elif key.__contains__('胸'):
                chest += key + ':' + value + ','
            elif key.__contains__('肛门') or key.__contains__('指诊'):
                digital_rectal += key + ':' + value + ','
            elif key.__contains__('外生殖器'):
                external_genital += key + ':' + value + ','
            elif key.__contains__('脊柱'):
                spinal += key + ':' + value + ','
            elif key.__contains__('四肢'):
                limb += key + ':' + value + ','
            elif key.__contains__('神经'):
                neurological += key + ':' + value + ','
            else:
                general_survey += key + ':' + value + ','
        patient_info['一般状况检查'] = general_survey
        patient_info['皮肤和粘膜检査'] = skin_and_mucosal
        patient_info['全身浅表淋巴结检查'] = node
        patient_info['头部检査'] = head
        patient_info['颈部检査'] = neck
        patient_info['腹部检査'] = abdomen
        patient_info['胸部检査'] = chest
        patient_info['肛门指诊检查'] = digital_rectal
        patient_info['外生殖器检査'] = external_genital
        patient_info['脊柱检查'] = spinal
        patient_info['四肢检查'] = limb
        patient_info['神经系统检查'] = neurological
        patient_info['中医四诊'] = chinese


def clean_string(s):
    """去除字符串中的空格和 <> 符号"""
    return s.replace(" ", "").replace("<", "&lt;").replace(">", "&gt;")

def clean_dict(d):
    """递归地清理字典中的所有键和值，并忽略 None 的键和值"""
    if isinstance(d, dict):
        cleaned_dict = {}
        for k, v in d.items():
            if k is not None:
                cleaned_key = clean_string(k)
                if v is not None:
                    cleaned_value = clean_dict(v)
                else:
                    cleaned_value = v
                cleaned_dict[cleaned_key] = cleaned_value
        return cleaned_dict
    elif isinstance(d, list):
        return [clean_dict(i) for i in d]
    elif isinstance(d, str):
        return clean_string(d)
    else:
        return d


def parse_admission_record(data):
    patient_info = parse_patient_document_by_str(data)
    # 将 Python 对象转换为格式化的 JSON 字符串
    # formatted_json = json.dumps(patient_info, indent=4, ensure_ascii=False)
    # print(formatted_json)
    merge_data(patient_info)
    patient_info = clean_dict(patient_info)
    cda_data = assembling_cda_record(patient_info, 1)
    return cda_data



# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    # # 遍历所有目录中xml 文件完成解析，并生成入院记录 cda 文档
    # directory = '/Users/gaoyanliang/nsyy/病历解析/入院记录/bingli'
    # for root, dirs, files in os.walk(directory):
    #     for file in files:
    #         if file.endswith(".xml"):
    #             patient_info = parse_patient_document(os.path.join(root, file))
    #             # 将 Python 对象转换为格式化的 JSON 字符串
    #             # formatted_json = json.dumps(patient_info, indent=4, ensure_ascii=False)
    #             # print(formatted_json)
    #             try:
    #                 merge_data(patient_info)
    #                 patient_info = clean_dict(patient_info)
    #                 assembling_cda_record(patient_info, 1)
    #             except Exception as e:
    #                 print(file, '===> assembling_admission_record 解析异常')

    # for s in sign_set:
    #     print(s)

    patient_info = parse_patient_document('/Users/gaoyanliang/nsyy/病历解析/入院记录/bingli/心内科/心内科三病区入院记录_2024-03-02_11-28-44.xml')
    merge_data(patient_info)
    patient_info = clean_dict(patient_info)
    # formatted_json = json.dumps(patient_info, indent=4, ensure_ascii=False)
    # print(formatted_json)
    assembling_cda_record(patient_info, 1)


    # write_to_excel(patient_info)










