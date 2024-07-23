import json
import xml.etree.ElementTree as ET
import re
import os

"""
入院记录解析
"""

section_info = {}


def parse_patient_document(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    patient_info = {}

    # 提取姓名，性别，年龄等基本信息
    patient_info['姓名'] = root.find(".//element[@title='姓名']").text

    # 性别 不规范 急诊外科入院记录_2024-03-02_17-22-47.xml
    # patient_info['性别'] = root.find(".//e_enum[@title='性别']").find('enumvalues/element').text
    patient_info['年龄'] = root.find(".//element[@title='年龄']").text
    patient_info['科室'] = root.find(".//element[@title='科室']").text
    if root.find(".//element[@title='职业']") is not None:
        patient_info['职业'] = root.find(".//element[@title='职业']").text

    # xml 中 身份证不规范， 普儿入院记录_2024-03-01_19-13-59.xml  普儿入院记录_2024-03-01_09-53-20.xml
    if root.find(".//element[@title='身份证件号码']") is not None:
        patient_info['身份证号'] = root.find(".//element[@title='身份证件号码']").text
    patient_info['入院时间'] = root.find(".//element[@title='入院日期']").text
    patient_info['病史采集时间'] = root.find(".//element[@title='信息录入日期时间']").text if root.find(".//element[@title='信息录入日期时间']") else root.find(".//element[@title='入院日期']").text
    # 病史叙述者 有可能为空
    if root.find(".//e_enum[@title='病史陈述者']") is not None and root.find(".//e_enum[@title='病史陈述者']").find('enumvalues/element') is not None:
        patient_info['病史叙述者'] = root.find(".//e_enum[@title='病史陈述者']").find('enumvalues/element').text
    else:
        patient_info['病史叙述者'] = ''

    header = root.find('./document')

    # 特殊处理
    for utext in header.iter('utext'):
        if utext.text is None:
            continue
        text = utext.text.replace(' ', '').replace(':', '').replace('：', '')
        if text == '与患者关系':
            utest_no = int(utext.get('no'))
            value_no = str(utest_no + 1)
            patient_info['与患者关系'] = root.find(".//utext[@no=" + "\'" + value_no + "\'" "]").text.strip(": ")
        if text == '职业' and '职业' not in patient_info:
            utest_no = int(utext.get('no'))
            value_no = str(utest_no + 1)
            patient_info['职业'] = root.find(".//utext[@no=" + "\'" + value_no + "\'" "]").text.strip(": ")


    for section in header.iter('section'):
        parse_hpi(section, patient_info, xml_file)

    return patient_info


def parse_enum(enum, rangexml: str = ''):
    if not rangexml:
        if enum.find('enumvalues'):
            rangexml = enum.find('enumvalues').tail if enum.find('enumvalues') else ''
        else:
            rangexml = enum.text if enum.text else ''
    if not rangexml:
        return ''
    # 使用正则表达式匹配并提取 </root>' 后面的汉字部分
    pattern = re.compile(r"rangexml='[^']*'\s*(\S+)")
    match = pattern.search(rangexml)
    if match:
        extracted_text = match.group(1)
        return extracted_text
    else:
        print(enum.get('title'), " ===> 未能提取出汉字部分")
    return ''


def skip_continue(father_sid, sid):
    if father_sid == 'AB115ABF651443FA847D6DB6BDB0153E' and sid == '1A49854542FC4788ADFD8BA0273F4E9A':
        # 婚育史 过滤部分无效数据
        return True
    if father_sid == '7F98425DF2AD41109727949ABE9B3773' and sid == 'D850CDB8B9DD490D87CC6F752131BCC7':
        # 初步诊断 过滤 “单机这里选择职称”
        return True
    if father_sid == '61E75BB1B4B346008DFBEC79BEDE630F' and (sid == '0C6CFAB22A2C4AADAA480DA199D738D1' or sid == '138B3F46D506445FAA0B38F35739E8AB'):
        # 专科情况 过滤 “报告卡编码” "有无"
        return True
    return False


# 解析病历
def parse_hpi(section, info_dict, file):
    try:
        sid = section.get('sid') if section.get('sid') else section.get('iid')
        title = section.get('title')
        value = ''
        value_dict = {}
        for child in section:
            tag = child.tag
            if tag == 'break' or (tag == 'utext' and (str(child.text).replace(' ', '') == ':' or str(child.text).replace(' ', '') == '：')):
                continue
            if tag == 'utext':
                if child.text and not child.text.__contains__('家属签字确认'):
                    value = value + child.text.strip() if child.text else ''
            elif tag == 'element':
                if skip_continue(section.get('sid'), child.get('sid')):
                   continue
                text = child.text if child.text else '/'
                if text.__contains__('textstyleno') and child.get('value'):
                    text = child.get('value')
                value = value + text
                # 婚育史 过滤部分无效数据
                child_sid = child.get('sid') if child.get('sid') else child.get('iid')
                value_dict[child_sid] = text if 'value' not in child.attrib or not 'unit' in child.attrib else child.get('value', '') + ' ' + child.get('unit', '')
            elif tag == 'e_enum' or tag == 'e_list':
                if len(child) < 1:
                    continue
                if skip_continue(section.get('sid'), child.get('sid')) or child.get('title').replace(' ', '') == '报告卡编码':
                   continue
                text = parse_enum(child)
                value = value + text
                # 婚育史 过滤部分无效数据
                child_sid = child.get('sid') if child.get('sid') else child.get('iid')
                value_dict[child_sid] = text
            elif tag == 'group':
                for group in child.findall('group'):
                    group_text = ''
                    for item in group:
                        if (item.tag == 'utext' or item.tag == 'element') and item.text:
                            group_text = group_text + item.text.replace("textstyleno=2 unitstr='", '')
                        elif item.tag == 'e_enum':
                            enumtext = parse_enum(item)
                            group_text = group_text + enumtext
                            item_sid = item.get('sid') if item.get('sid') else item.get('iid')
                            if item_sid:
                                value_dict[item_sid] = enumtext
                    value = value + group_text
                    child_sid = child.get('sid') if child.get('sid') else child.get('iid')
                    value_dict[child_sid] = group_text
            elif tag in ('tab', 'image', 'patisign', 'table', 'signature'):
                # 暂时不处理，没意义
                continue
            else:
                print('===> 未处理 ', title, tag)

        if len(value_dict) < 2:
            info_dict[sid] = value.strip()
        else:
            value_dict['value'] = value.strip()
            info_dict[sid] = value_dict
    except Exception as e:
        print(file, '===> parse_hpi 解析异常', e)


def write_to_excel(data):
    import pandas as pd
    from openpyxl import load_workbook

    # 假设你已有的Excel文件名为'existing_file.xlsx'
    # file_path = '/Users/gaoyanliang/nsyy/病历解析/入院记录/parse_data.xlsx'
    file_path = 'parsed_data.xlsx'
    # 将字典转换为DataFrame
    df = pd.DataFrame(data)
    # 尝试加载现有的Excel文件
    try:
        # 使用openpyxl加载工作簿
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        writer.book = book

        # 检查是否已有数据表
        if 'Sheet1' in book.sheetnames:
            existing_df = pd.read_excel(file_path, sheet_name='Sheet1', engine='openpyxl')
            combined_df = pd.concat([existing_df, df], ignore_index=True)
        else:
            combined_df = df

        # 将合并后的数据写入Excel文件
        combined_df.to_excel(writer, index=False, sheet_name='Sheet1')
        writer.save()
        writer.close()
    except FileNotFoundError:
        # 如果文件不存在，创建新文件
        df.to_excel(file_path, index=False, engine='openpyxl')




section_info = {}
sid_info = {}
item_info = set()
exception_sid = {}
all_sid_dict = {}
sid_set = set()
# 验证 document - element 中的 sid 是否统一
def document_section(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    header = root.find('./document')

    for section in header.findall('section'):
        section_sid = section.get('sid') if section.get('sid') else section.get('iid')
        title = section.get('title').replace(' ', '')

        section_info[title] = section_sid
        sid_info[section_sid] = title

        section_sid_set = set()
        if section_sid in all_sid_dict:
            sid_dict = all_sid_dict[section_sid]
        else:
            sid_dict = {'sid': section_sid, 'title': title, 'child': {}}
        for item in section:
            if ('sid' in item.attrib or 'iid' in item.attrib) and 'title' in item.attrib:
                sid = item.get('sid') if item.get('sid') else item.get('iid')
                title = item.get('title').replace(' ', '')

                if sid is None or title.__contains__('单击这里选择职称') or title.__contains__('报告卡编码') or item.tag == 'patisign' or title.__contains__('输入内容'):
                    continue

                if title.__contains__('身高数值') or title.__contains__('体重'):
                    print(xml_file)

                if sid in sid_info and sid_info[sid] != title:
                    exception_sid[sid] = { "sid": sid, "title1": title, "title2": sid_info[sid]}

                if sid in section_sid_set:
                    continue
                section_sid_set.add(sid)
                section_info[title] = sid
                sid_info[sid] = title
                sid_dict['child'][sid] = {'sid': sid, 'title': title, 'child': {}}

                if item.tag == 'group':
                    for group in item.findall('group'):
                        for ite in group:
                            if ite.tag == 'e_enum':
                                group_sid = ite.get('sid') if ite.get('sid') else ite.get('iid')
                                group_title = ite.get('title').replace(' ', '')

                                if group_sid is None or group_title.__contains__('单击这里选择职称') or group_title.__contains__(
                                        '报告卡编码') or ite.tag == 'patisign':
                                    continue

                                if group_sid in sid_info and sid_info[group_sid] != group_title:
                                    exception_sid[group_sid] = {"sid": group_sid, "title1": group_title, "title2": sid_info[group_sid]}

                                if group_sid in section_sid_set:
                                    continue
                                section_sid_set.add(group_sid)
                                section_info[group_title] = group_sid
                                sid_info[group_sid] = group_title
                                sid_dict['child'][sid]['child'][group_sid] = {'sid': group_sid, 'title': group_title}

        all_sid_dict[section_sid] = sid_dict


def load_local_json():
    # JSON文件路径
    file_path = 'all_sid.json'
    # 打开文件并加载JSON数据
    with open(file_path, 'r') as file:
        data = json.load(file)
    # 打印加载的数据
    print(data)


# Press the green button in the gutter to run the script.
if __name__ == '__main__':

    # import xml.etree.ElementTree as ET
    # # 解析XML文件
    # tree = ET.parse('/Users/gaoyanliang/nsyy/病历解析/入院记录/入院记录文档示例.xml')
    # root = tree.getroot()
    # print()

    # directory = '/Users/gaoyanliang/nsyy/病历解析/入院记录/bingli/白内障/'
    # for root, dirs, files in os.walk(directory):
    #     for file in files:
    #         if file.endswith(".xml"):
    #             patient_info = parse_patient_document(os.path.join(root, file))
    #             # 将 Python 对象转换为格式化的 JSON 字符串
    #             formatted_json = json.dumps(patient_info, indent=4, ensure_ascii=False)
    #             print(formatted_json)


    patient_info = parse_patient_document('/Users/gaoyanliang/nsyy/病历解析/入院记录/bingli/普儿/普儿入院记录_2024-03-01_09-53-20.xml')
    formatted_json = json.dumps(patient_info, indent=4, ensure_ascii=False)
    print(formatted_json)
    #
    # write_to_excel(patient_info)








