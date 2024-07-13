import xml.etree.ElementTree as ET
import re
import os

"""
入院记录解析
"""

miss_list = []

section_info = {}

def parse_patient_document(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    patient_info = {}

    # 提取姓名，性别，年龄等基本信息
    patient_info['姓名'] = root.find(".//element[@title='姓名']").text

    # 性别 不规范 急诊外科入院记录_2024-03-02_17-22-47.xml
    # patient_info['性别'] = root.find(".//e_enum[@title='性别']").find('enumvalues/element').text
    patient_info['年龄'] = root.find(".//element[@title='年龄']").text
    patient_info['科室'] = root.find(".//element[@title='科室']").text
    if root.find(".//element[@title='职业']"):
        patient_info['职业'] = root.find(".//element[@title='职业']").text

    # xml 中 身份证不规范， 普儿入院记录_2024-03-01_19-13-59.xml  普儿入院记录_2024-03-01_09-53-20.xml
    if root.find(".//element[@title='身份证件号码']"):
        patient_info['身份证号'] = root.find(".//element[@title='身份证件号码']").text
    patient_info['入院时间'] = root.find(".//element[@title='入院日期']").text
    patient_info['病史采集时间'] = root.find(".//element[@title='信息录入日期时间']").text if root.find(".//element[@title='信息录入日期时间']") else root.find(".//element[@title='入院日期']").text
    # 病史叙述者 有可能为空
    if root.find(".//e_enum[@title='病史陈述者']") and root.find(".//e_enum[@title='病史陈述者']").find('enumvalues/element'):
        patient_info['病史叙述者'] = root.find(".//e_enum[@title='病史陈述者']").find('enumvalues/element').text
    else:
        patient_info['病史叙述者'] = ''

    header = root.find('./document')

    # 特殊处理
    for utext in header.iter('utext'):
        if utext.text is None:
            continue
        text = utext.text.replace(' ', '').replace(':', '').replace('：', '')
        if text == '与患者关系':
            utest_no = int(utext.get('no'))
            value_no = str(utest_no + 1)
            patient_info['与患者关系'] = root.find(".//utext[@no=" + "\'" + value_no + "\'" "]").text.strip(": ")
        if text == '职业' and '职业' not in patient_info:
            utest_no = int(utext.get('no'))
            value_no = str(utest_no + 1)
            patient_info['职业'] = root.find(".//utext[@no=" + "\'" + value_no + "\'" "]").text.strip(": ")


    for section in header.iter('section'):
        title = section.get('title')
        title = title.replace(' ', '')

        # if title in ('主诉', '现病史', '既往史', '个人史', '婚育史', '家族史', '体格检查', '初步诊断', '专科情况', '辅助检查', '月经史'):
        #     parse(section, patient_info)
        # elif title in ('病人信息', '住院医师', '签名'):
        #     continue
        # else:
        #     miss_list.append(title)

        if title == '主诉':
            parse(section, patient_info)
            parse_chief_complaint(section, patient_info)
        elif title == '现病史':
            parse_hpi(section, patient_info)
        elif title == '既往史':
            parse_past_medical_history(section, patient_info)
        elif title == '个人史':
            parse_personal_history(section, patient_info)
        elif title in ('婚育史', '外阴', '阴道', '宫颈', '宫体', '附件'):
            parse_marital_history(section, patient_info)
        elif title in ('家族史', '母孕史'):
            parse_family_history(section, patient_info)
        elif title == '体格检查':
            parse_physical_examination(section, patient_info)
        elif title in ('初步诊断', '专科情况', '辅助检查', '月经史', '神色形态', '声音', '气味', '舌象', '脉象', '初步诊断', '步诊断',
                       '中医诊断', '证型', '西医诊断', '专科检查', '职业史', '心理史', '康复评定', 'NIHSS评分', 'Hunt-Hess评分',
                       'Fugl-Meyer平衡功能测定', '改良巴氏指数评定', '脑卒中Brunnstrom分期'):
            # 神经外科入院记录_2024-03-01_12-35-07.xml   神经外科入院记录_2024-03-01_21-52-03.xml  {'步诊断'}
            parse_special_situation(section, patient_info)
        elif title in ('病人信息', '住院医师', '签名'):
            continue
        else:
            miss_list.append(title)

    return patient_info


# 通用解析逻辑
def parse(section, info_dict):
    try:
        if section.get('sid') not in section_info:
            section_info[section.get('sid')] = section.attrib
        title = section.get('title')
        value = ''
        for child in section:
            if child.tag == 'break' or (child.tag == 'utext' and (str(child.text).replace(' ', '') == ':' or str(child.text).replace(' ', '') == '：')):
                continue
            if (child.tag == 'utext' or child.tag == 'element') and child.text and not child.text.__contains__('家属签字确认'):
                value = value + child.text
            if child.tag == 'e_enum' or child.tag == 'e_list':
                if len(child) < 1 and (not child.find('enumvalues') or not child.find('enumvalues').tail):
                    continue
                enum_title = child.get('title')
                # 使用正则表达式匹配并提取 </root>' 后面的汉字部分
                pattern = re.compile(r"rangexml='[^']*'\s*(\S+)")
                match = pattern.search(child.find('enumvalues').tail)
                if match:
                    extracted_text = match.group(1)
                    value = value + enum_title + extracted_text
                else:
                    print(title, enum_title, "未能提取出汉字部分")
            if child.tag == 'group':
                for group in child.findall('group'):
                    value = value + child.get('title') + ':'
                    for item in group:
                        if item.tag == 'utext' and item.text:
                            value = value + item.text
                        if item.tag == 'element':
                            value = value + item.text if item.text else item.get('value')
                        if item.tag == 'e_enum':
                            if not item.find('enumvalues') or not item.find('enumvalues').tail:
                                continue
                            item_title = item.get('title')
                            # 使用正则表达式匹配并提取 </root>' 后面的汉字部分
                            pattern = re.compile(r"rangexml='[^']*'\s*(\S+)")
                            match = pattern.search(item.find('enumvalues').tail)
                            if match:
                                extracted_text = match.group(1)
                                value = value + item_title + extracted_text
                            else:
                                print(title, item_title, " 未能提取出汉字部分")
        info_dict[title] = value
    except Exception as e:
        print(title, '解析异常', e)
        info_dict[title] = '解析异常'


# 解析主诉
def parse_chief_complaint(section, info_dict):
    try:
        if section.get('sid') not in section_info:
            section_info[section.get('sid')] = section.attrib
        title = section.get('title')
        value = ''
        for child in section:
            if child.tag == 'break' or (child.tag == 'utext' and (str(child.text).replace(' ', '') == ':' or str(child.text).replace(' ', '') == '：')):
                continue
            if child.tag == 'utext' and child.text:
                value = value + child.text
        info_dict[title] = value
    except Exception as e:
        print('主诉解析异常', e)
        info_dict['主诉'] = '主诉解析异常'


# 解析现病史
def parse_hpi(section, info_dict):
    try:
        if section.get('sid') not in section_info:
            section_info[section.get('sid')] = section.attrib
        title = section.get('title')
        value = ''
        for child in section:
            if child.tag == 'break' or (child.tag == 'utext' and (str(child.text).replace(' ', '') == ':' or str(child.text).replace(' ', '') == '：')):
                continue
            if (child.tag == 'utext' or child.tag == 'element') and child.text:
                value = value + child.text
            if child.tag == 'group':
                for group in child.findall('group'):
                    for item in group:
                        if (item.tag == 'utext' or item.tag == 'element') and item.text:
                            value = value + item.text
                        elif item.tag == 'e_enum':
                            item_title = item.get('title')
                            # 使用正则表达式匹配并提取 </root>' 后面的汉字部分
                            pattern = re.compile(r"rangexml='[^']*'\s*(\S+)")
                            match = pattern.search(item.find('enumvalues').tail)
                            if match:
                                extracted_text = match.group(1)
                                value = value + item_title + extracted_text
                            else:
                                print("现病史 ", item_title, " 未能提取出汉字部分")
        info_dict[title] = value
    except Exception as e:
        print('现病史解析异常', e)
        info_dict['现病史'] = '现病史解析异常'


# 解析既往史
def parse_past_medical_history(section, info_dict):
    try:
        if section.get('sid') not in section_info:
            section_info[section.get('sid')] = section.attrib
        title = section.get('title')
        value = ''
        for child in section:
            if child.tag == 'break' or (child.tag == 'utext' and (str(child.text).replace(' ', '') == ':' or str(child.text).replace(' ', '') == '：')):
                continue
            if (child.tag == 'utext' or child.tag == 'element') and child.text:
                value = value + child.text
            if child.tag == 'e_enum' and child.attrib and child.attrib.get('title') == '健康状况':
                # 使用正则表达式匹配并提取 </root>' 后面的汉字部分
                pattern = re.compile(r"rangexml='[^']*'\s*(\S+)")
                match = pattern.search(child.find('enumvalues').tail)
                if match:
                    extracted_text = match.group(1)
                    value = value + child.attrib.get('title') + extracted_text
                else:
                    print("既往史 健康状况 未能提取出汉字部分")

            if child.tag == 'group':
                for group in child.findall('group'):
                    value = value + child.get('title') + ':'
                    group_value = ''
                    for item in group:
                        if item.tag == 'utext' and item.text:
                            group_value = group_value + item.text
                        if item.tag == 'element':
                            group_value = group_value + item.text if item.text else item.get('value')
                        if item.tag == 'e_enum':
                            if not item.text:
                                continue
                            # 使用正则表达式匹配并提取 </root>' 后面的汉字部分
                            pattern = re.compile(r"rangexml='[^']*'\s*(\S+)")
                            match = pattern.search(item.text)
                            if match:
                                extracted_text = match.group(1)
                                group_value = group_value + extracted_text
                            else:
                                print("未能提取出汉字部分")

                    info_dict[child.get('title')] = group_value
                    value = value + group_value

        info_dict[title] = value
    except Exception as e:
        print('既往史解析异常', e)
        info_dict['既往史'] = '既往史解析异常'


# 解析个人史
def parse_personal_history(section, info_dict):
    try:
        if section.get('sid') not in section_info:
            section_info[section.get('sid')] = section.attrib
        title = section.get('title')
        value = ''
        for child in section:
            if child.tag == 'break' or (child.tag == 'utext' and (str(child.text).replace(' ', '') == ':' or str(child.text).replace(' ', '') == '：')):
                continue
            if child.tag == 'utext' and child.text:
                value = value + child.text
            if child.tag == 'element' and child.text:
                value = value + child.get('title') + ':' + child.text
            if child.tag == 'group':
                for group in child.findall('group'):
                    value = value + group.get('title') + ':'
                    for item in group:
                        if (item.tag == 'utext' or item.tag == 'element') and item.text:
                            value = value + item.text

        info_dict[title] = value
    except Exception as e:
        print('个人史解析异常', e)
        info_dict['个人史'] = '个人史解析异常'


# 解析婚育史
def parse_marital_history(section, info_dict):
    try:
        if section.get('sid') not in section_info:
            section_info[section.get('sid')] = section.attrib
        title = section.get('title')
        value = ''
        for child in section:
            if child.tag == 'break' or (child.tag == 'utext' and (str(child.text).replace(' ', '') == ':' or str(child.text).replace(' ', '') == '：')):
                continue
            if child.tag == 'utext' and child.text:
                value = value + child.text
            if child.tag == 'element' and child.text:
                value = value + child.get('title') + ':' + child.text
            if child.tag == 'group':
                for group in child.findall('group'):
                    for item in group:
                        if (item.tag == 'element' or item.tag == 'utext') and item.text:
                            value = value + item.text
                        elif item.tag == 'e_enum':
                            item_title = item.get('title')
                            # 使用正则表达式匹配并提取 </root>' 后面的汉字部分
                            pattern = re.compile(r"rangexml='[^']*'\s*(\S+)")
                            match = pattern.search(item.find('enumvalues').tail)
                            if match:
                                extracted_text = match.group(1)
                                value = value + item_title + extracted_text
                            else:
                                print("婚育史 ", item_title, " 未能提取出汉字部分")

        info_dict[title] = value
    except Exception as e:
        print('婚育史解析异常', e)
        info_dict['婚育史'] = '婚育史解析异常'


# 解析家族史
def parse_family_history(section, info_dict):
    try:
        if section.get('sid') not in section_info:
            section_info[section.get('sid')] = section.attrib
        title = section.get('title')
        value = ''
        for child in section:
            if child.tag == 'break' or (child.tag == 'utext' and (str(child.text).replace(' ', '') == ':' or str(child.text).replace(' ', '') == '：')):
                continue
            if (child.tag == 'utext' or child.tag == 'element') and child.text and not child.text.__contains__('家属签字确认'):
                value = value + child.text
            if child.tag == 'e_enum':
                item_title = child.get('title')
                # 使用正则表达式匹配并提取 </root>' 后面的汉字部分
                pattern = re.compile(r"rangexml='[^']*'\s*(\S+)")
                match = pattern.search(child.find('enumvalues').tail)
                if match:
                    extracted_text = match.group(1)
                    value = value + item_title + extracted_text
                else:
                    print("家族史 ", item_title, " 未能提取出汉字部分")
            if child.tag == 'group':
                for group in child.findall('group'):
                    for item in group:
                        if item.tag == 'utext':
                            value = value + item.text
                        if item.tag == 'element':
                            value = value + item.get('value')

        info_dict[title] = value
    except Exception as e:
        print('家族史解析异常', e)
        info_dict['家族史'] = '家族史解析异常'


# 解析体格检查
def parse_physical_examination(section, info_dict):
    try:
        if section.get('sid') not in section_info:
            section_info[section.get('sid')] = section.attrib
        title = section.get('title')
        value = ''
        for child in section:
            if child.tag == 'break' or (child.tag == 'utext' and (str(child.text).replace(' ', '') == ':' or str(child.text).replace(' ', '') == '：')):
                continue
            if (child.tag == 'utext' or child.tag == 'element') and child.text and not child.text.__contains__('家属签字确认'):
                value = value + child.text + ' '
            if child.tag == 'e_enum' or child.tag == 'e_list':
                if len(child) < 1:
                    continue
                item_title = child.get('title')
                # 使用正则表达式匹配并提取 </root>' 后面的汉字部分
                pattern = re.compile(r"rangexml='[^']*'\s*(\S+)")
                if child.find('enumvalues'):
                    match = pattern.search(child.find('enumvalues').tail)
                else:
                    match = pattern.search(child.text)
                if match:
                    extracted_text = match.group(1)
                    value = value + extracted_text
                else:
                    print("体格检查 ", item_title, " 未能提取出汉字部分")
            if child.tag == 'group':
                for group in child.findall('group'):
                    for item in group:
                        if item.tag == 'utext':
                            value = value + item.text

        info_dict[title] = value
    except Exception as e:
        print('体格检查解析异常', e)
        info_dict['体格检查'] = '体格检查解析异常'


# 解析专科情况
def parse_special_situation(section, info_dict):
    try:
        if section.get('sid') not in section_info:
            section_info[section.get('sid')] = section.attrib
        title = section.get('title')
        value = ''
        for child in section:
            if child.tag == 'break' or (child.tag == 'utext' and (str(child.text).replace(' ', '') == ':' or str(child.text).replace(' ', '') == '：')):
                continue
            if (child.tag == 'utext' or child.tag == 'element' or child.tag == 'e_enum') and child.text and not child.text.__contains__('家属签字确认'):
                value = value + child.text + ' '
        info_dict[title] = value
    except Exception as e:
        print('专科情况/辅助检查解析异常', e)
        info_dict['专科情况/辅助检查'] = '专科情况/辅助检查解析异常'


def write_to_excel(data):
    import pandas as pd
    from openpyxl import load_workbook

    # 假设你已有的Excel文件名为'existing_file.xlsx'
    # file_path = '/Users/gaoyanliang/nsyy/病历解析/入院记录/parse_data.xlsx'
    file_path = 'existing_file.xlsx'
    # 将字典转换为DataFrame
    df = pd.DataFrame(data)
    # 尝试加载现有的Excel文件
    try:
        # 使用openpyxl加载工作簿
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        writer.book = book

        # 检查是否已有数据表
        if 'Sheet1' in book.sheetnames:
            # 读取现有数据表
            existing_df = pd.read_excel(file_path, sheet_name='Sheet1', engine='openpyxl')
            # 将新数据追加到现有数据
            combined_df = pd.concat([existing_df, df], ignore_index=True)
        else:
            combined_df = df

        # 将合并后的数据写入Excel文件
        combined_df.to_excel(writer, index=False, sheet_name='Sheet1')
        writer.save()
        writer.close()
    except FileNotFoundError:
        # 如果文件不存在，创建新文件
        df.to_excel(file_path, index=False, engine='openpyxl')

    print("Data appended successfully!")


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    patient_info = parse_patient_document("/Users/gaoyanliang/nsyy/病历解析/入院记录/bingli/眼科/眼科入院记录_2024-03-02_18-52-04.xml")
    for key, value in patient_info.items():
        print(f"{key}:  {value}")

    # patient_info_list = []
    # directory = '/Users/gaoyanliang/nsyy/病历解析/入院记录/bingli/'
    # for root, dirs, files in os.walk(directory):
    #     for file in files:
    #         if file.endswith(".xml"):
    #             miss_list = []
    #             try:
    #                 # print('开始解析 ', file)
    #                 patient_info = parse_patient_document(os.path.join(root, file))
    #                 patient_info['文件来源'] = file
    #                 # for key, value in patient_info.items():
    #                 #     print(f"{key}:  {value}")
    #                 patient_info_list.append(patient_info)
    #             except Exception as e:
    #                 print('解析文件 ', file, ' 异常', e)
    #             # print()
    #             if miss_list:
    #                 print(file, '缺失', set(miss_list))

    # 将解析的数据写入Excel文件
    # write_to_excel(patient_info_list)




